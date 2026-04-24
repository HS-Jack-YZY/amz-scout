"""Smoke tests for the webapp scaffold (Phase 1).

These tests do NOT hit the real Anthropic API or the real Chainlit server.
They verify import integrity, tool dispatch envelope shape, and auth
callback behavior in isolation.
"""

import asyncio
import sys

import pytest


def _reset_webapp_modules() -> None:
    """Clear any cached webapp.* imports so env changes take effect."""
    for mod in list(sys.modules):
        if mod.startswith("webapp"):
            del sys.modules[mod]


def _set_fake_env(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("ANTHROPIC_API_KEY", "fake")
    monkeypatch.setenv("CHAINLIT_AUTH_SECRET", "fake")
    monkeypatch.setenv("APP_PASSWORD", "fake")
    monkeypatch.setenv("KEEPA_API_KEY", "fake")


@pytest.mark.unit
class TestWebappImports:
    """Verify the webapp module imports cleanly with minimal env."""

    def test_config_imports(self, monkeypatch: pytest.MonkeyPatch) -> None:
        _set_fake_env(monkeypatch)
        _reset_webapp_modules()
        from webapp import config

        config.validate_env()  # should not raise
        assert config.MODEL_ID == "claude-sonnet-4-6"
        assert config.ALLOWED_EMAIL_DOMAIN == "@gl-inet.com"

    def test_validate_env_raises_when_missing(self, monkeypatch: pytest.MonkeyPatch) -> None:
        # Set to empty string (not delete) so dotenv's default override=False
        # won't repopulate from an existing .env file when webapp.config re-imports.
        for var in ("ANTHROPIC_API_KEY", "CHAINLIT_AUTH_SECRET", "APP_PASSWORD", "KEEPA_API_KEY"):
            monkeypatch.setenv(var, "")
        _reset_webapp_modules()
        from webapp import config

        with pytest.raises(ValueError, match="Missing required environment variables"):
            config.validate_env()


@pytest.mark.unit
class TestEmailDomainNormalization:
    """Regression guard for the auth domain anchor.

    ``webapp.config.ALLOWED_EMAIL_DOMAIN`` MUST always start with "@" and be
    lowercased, regardless of how the operator wrote the env var. Without
    this normalization, an env value of "gl-inet.com" (no leading "@") would
    let "attacker@evilgl-inet.com" satisfy ``email.endswith(domain)`` and
    bypass the email whitelist entirely.
    """

    def test_default_value_already_anchored(self, monkeypatch: pytest.MonkeyPatch) -> None:
        _set_fake_env(monkeypatch)
        monkeypatch.delenv("ALLOWED_EMAIL_DOMAIN", raising=False)
        _reset_webapp_modules()
        from webapp import config

        assert config.ALLOWED_EMAIL_DOMAIN == "@gl-inet.com"

    def test_missing_at_prefix_is_added(self, monkeypatch: pytest.MonkeyPatch) -> None:
        _set_fake_env(monkeypatch)
        monkeypatch.setenv("ALLOWED_EMAIL_DOMAIN", "gl-inet.com")
        _reset_webapp_modules()
        from webapp import config

        assert config.ALLOWED_EMAIL_DOMAIN == "@gl-inet.com"

    def test_uppercase_input_is_lowercased(self, monkeypatch: pytest.MonkeyPatch) -> None:
        _set_fake_env(monkeypatch)
        monkeypatch.setenv("ALLOWED_EMAIL_DOMAIN", "GL-INET.COM")
        _reset_webapp_modules()
        from webapp import config

        assert config.ALLOWED_EMAIL_DOMAIN == "@gl-inet.com"

    def test_lookalike_domain_does_not_satisfy_endswith(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """The whole point of the anchor: a normalized domain must reject
        ``attacker@evilgl-inet.com``. We verify by hand-rolling the same
        ``endswith`` check ``webapp.auth`` uses, against the normalized value.
        """
        _set_fake_env(monkeypatch)
        monkeypatch.setenv("ALLOWED_EMAIL_DOMAIN", "gl-inet.com")  # no "@"
        _reset_webapp_modules()
        from webapp import config

        normalized = config.ALLOWED_EMAIL_DOMAIN
        assert not "attacker@evilgl-inet.com".endswith(normalized)
        assert "alice@gl-inet.com".endswith(normalized)

    def test_empty_domain_falls_back_to_default(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """An empty ALLOWED_EMAIL_DOMAIN must fall back to the default
        rather than silently producing '@' which locks out all users.
        """
        _set_fake_env(monkeypatch)
        monkeypatch.setenv("ALLOWED_EMAIL_DOMAIN", "")
        _reset_webapp_modules()
        from webapp import config

        assert config.ALLOWED_EMAIL_DOMAIN == "@gl-inet.com"

    def test_whitespace_only_domain_falls_back(self, monkeypatch: pytest.MonkeyPatch) -> None:
        _set_fake_env(monkeypatch)
        monkeypatch.setenv("ALLOWED_EMAIL_DOMAIN", "   ")
        _reset_webapp_modules()
        from webapp import config

        assert config.ALLOWED_EMAIL_DOMAIN == "@gl-inet.com"


@pytest.mark.unit
class TestToolDispatch:
    """Verify tool dispatch returns the amz_scout.api envelope shape."""

    def test_unknown_tool_returns_envelope(self, monkeypatch: pytest.MonkeyPatch) -> None:
        _set_fake_env(monkeypatch)
        _reset_webapp_modules()
        from webapp.tools import dispatch_tool

        result = asyncio.run(dispatch_tool("nonexistent", {}))
        assert result["ok"] is False
        assert "Unknown tool" in result["error"]
        assert "data" in result
        assert "meta" in result

    def test_tool_schemas_have_cache_control_on_last(self, monkeypatch: pytest.MonkeyPatch) -> None:
        _set_fake_env(monkeypatch)
        _reset_webapp_modules()
        from webapp.tools import TOOL_SCHEMAS

        assert len(TOOL_SCHEMAS) >= 1
        assert "cache_control" in TOOL_SCHEMAS[-1], (
            "The last tool must have cache_control for prompt caching to work"
        )
        # Only the last tool should carry cache_control; earlier tools must not.
        for tool in TOOL_SCHEMAS[:-1]:
            assert "cache_control" not in tool, (
                "Only the LAST tool should have cache_control; it caches all preceding tools"
            )

    def test_all_expected_tools_present(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Every shipped tool must appear in TOOL_SCHEMAS.

        Server-side tools (Anthropic-executed, e.g. ``web_search``) carry both
        ``type`` and ``name`` — they share the ``name`` namespace with
        client-side tools but are dispatched by Anthropic, not our
        ``dispatch_tool``. The expected set covers BOTH kinds.
        """
        _set_fake_env(monkeypatch)
        _reset_webapp_modules()
        from webapp.tools import TOOL_SCHEMAS

        names = {tool["name"] for tool in TOOL_SCHEMAS}
        expected = {
            "query_latest",
            "check_freshness",
            "keepa_budget",
            "query_availability",
            "query_compare",
            "query_deals",
            "query_ranking",
            "query_sellers",
            "query_trends",
            "web_search",
            "register_asin_from_url",
        }
        assert names == expected, f"Missing or extra tools: {names ^ expected}"

    def test_dispatcher_routes_all_known_tools(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Every declared tool must route through dispatch_tool without KeyError/AttributeError.

        This test exercises ONLY the dispatcher — it does not want to hit the real
        amz_scout.api (which would touch the SQLite registry and, for query_trends
        / query_sellers / query_deals, attempt a real Keepa network call). So we:

        1. Patch `cl.step` to a no-op decorator BEFORE importing webapp.tools, since
           the real decorator requires an active Chainlit session context.
        2. Patch every `_api_*` alias on the freshly-imported webapp.tools module to
           a fake that returns a well-formed envelope dict. This bypasses the real
           API entirely and asserts only dispatcher routing + envelope shape.
        """
        _set_fake_env(monkeypatch)

        import chainlit as cl

        def _noop_step(**_kwargs):  # type: ignore[no-untyped-def]
            def _decorator(fn):  # type: ignore[no-untyped-def]
                return fn

            return _decorator

        monkeypatch.setattr(cl, "step", _noop_step)

        _reset_webapp_modules()
        from webapp import tools as webapp_tools
        from webapp.tools import TOOL_SCHEMAS, dispatch_tool

        def _fake_envelope(*_args, **_kwargs) -> dict:  # type: ignore[no-untyped-def]
            return {"ok": True, "data": [], "error": None, "meta": {"stub": True}}

        for attr in (
            "_api_check_freshness",
            "_api_keepa_budget",
            "_api_query_availability",
            "_api_query_compare",
            "_api_query_deals",
            "_api_query_latest",
            "_api_query_ranking",
            "_api_query_sellers",
            "_api_query_trends",
            "_api_register_asin_from_url",
        ):
            monkeypatch.setattr(webapp_tools, attr, _fake_envelope)

        async def _run_all() -> list[tuple[str, dict]]:
            """Run every schema's dispatch in a single event loop.

            Collapsing to one `asyncio.run` (vs one per tool) keeps the test cheap
            and avoids masking event-loop-scoped bugs in future real wrappers.
            """
            out: list[tuple[str, dict]] = []
            for tool in TOOL_SCHEMAS:
                # Server-side tools (e.g. web_search, type="web_search_20260209")
                # don't declare an input_schema — Anthropic dispatches them on
                # its side. Skip: our dispatch_tool never sees them.
                if "input_schema" not in tool:
                    continue
                name = tool["name"]
                # Build minimal args dict satisfying required fields with safe placeholders
                args: dict = {}
                for prop in tool["input_schema"].get("required", []):
                    args[prop] = "UK" if prop == "marketplace" else "Slate 7"
                out.append((name, await dispatch_tool(name, args)))
            return out

        for name, result in asyncio.run(_run_all()):
            assert isinstance(result, dict), f"{name}: not a dict"
            assert "ok" in result, f"{name}: missing 'ok' key"
            assert "data" in result, f"{name}: missing 'data' key"
            assert "error" in result, f"{name}: missing 'error' key"
            assert "meta" in result, f"{name}: missing 'meta' key"

    def test_dispatcher_returns_error_envelope_when_required_field_missing(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Tools with required schema fields must return a clear validation envelope
        when the LLM drops the field, rather than passing an empty string to the API
        and getting a cryptic downstream resolution error.

        No Chainlit-context monkeypatching needed: the null-check returns BEFORE the
        `@cl.step`-decorated wrapper is ever called.
        """
        _set_fake_env(monkeypatch)
        _reset_webapp_modules()
        from webapp.tools import dispatch_tool

        tools_with_required = [
            ("query_latest", "marketplace"),
            ("query_compare", "product"),
            ("query_ranking", "marketplace"),
            ("query_sellers", "product"),
            ("query_trends", "product"),
            # register_asin_from_url validates brand first, so with empty args
            # the dispatcher returns on "brand". The other 3 required fields
            # (model/marketplace/amazon_url) share the same validation path.
            ("register_asin_from_url", "brand"),
        ]

        async def _run_all() -> list[tuple[str, str, dict]]:
            out: list[tuple[str, str, dict]] = []
            for tool_name, missing_field in tools_with_required:
                out.append((tool_name, missing_field, await dispatch_tool(tool_name, {})))
            return out

        for tool_name, missing_field, result in asyncio.run(_run_all()):
            assert result["ok"] is False, f"{tool_name}: expected ok=False, got {result}"
            assert missing_field in result["error"], (
                f"{tool_name}: error should mention '{missing_field}', got: {result['error']}"
            )
            assert "required" in result["error"], (
                f"{tool_name}: error should say 'required', got: {result['error']}"
            )
            assert result["data"] == []
            assert result["meta"] == {}


@pytest.mark.unit
class TestQueryPassthrough:
    """Phase 3 contract: row-emitting tools return summaries (not rows) to the
    LLM, and attach full DB rows as ``cl.File`` entries on the session.

    Supersedes the Phase 2 ``TestWebappTrimBoundary`` class whose row-shape
    assertions (``data`` is a list of trimmed rows) no longer hold — ``data``
    is now a summary dict. The regression surface is preserved here:
    ``test_full_rows_land_in_xlsx_not_trimmed`` guards against the trim
    decorator leaking back into the xlsx payload, and
    ``test_failure_envelope_passes_through`` keeps the failed-envelope
    passthrough rule.
    """

    def _patch_session_and_step(self, monkeypatch: pytest.MonkeyPatch) -> dict:
        """Monkey-patch the three chainlit surfaces touched by Phase 3 wiring.

        - ``cl.user_session`` → in-memory ``_FakeSession`` so we can assert on
          ``pending_files`` / ``query_log`` without Chainlit running.
        - ``cl.step`` → no-op decorator (same as TestToolDispatch).
        - ``cl.File`` → dataclass stub. ``cl.File.__post_init__`` reads
          ``context.session.thread_id`` and raises ``AttributeError`` when the
          session is None (pytest, CLI, background threads), so ``cl.File(...)``
          construction in ``_attach_file_to_session`` would explode inside its
          try/except and silently swallow the attach. Stubbing keeps the
          attach observable from the test.
        """
        from dataclasses import dataclass

        import chainlit as cl

        store: dict = {}

        class _FakeSession:
            def get(self, k, default=None):
                return store.get(k, default)

            def set(self, k, v):
                store[k] = v

        @dataclass
        class _FakeFile:
            name: str = ""
            content: bytes | str | None = None
            mime: str | None = None
            display: str = "inline"

        def _noop_step(**_kwargs):  # type: ignore[no-untyped-def]
            def _decorator(fn):  # type: ignore[no-untyped-def]
                return fn

            return _decorator

        monkeypatch.setattr(cl, "user_session", _FakeSession())
        monkeypatch.setattr(cl, "step", _noop_step)
        monkeypatch.setattr(cl, "File", _FakeFile)
        return store

    def test_query_trends_returns_summary_not_rows(self, monkeypatch: pytest.MonkeyPatch) -> None:
        _set_fake_env(monkeypatch)
        self._patch_session_and_step(monkeypatch)
        _reset_webapp_modules()
        from webapp import tools as webapp_tools
        from webapp.tools import dispatch_tool

        rows = [
            {"date": f"2026-04-{i:02d} 10:00", "value": 100 + i, "keepa_ts": 7584000 + i}
            for i in range(1, 88)
        ]

        def _fake(**_kw) -> dict:  # type: ignore[no-untyped-def]
            return {
                "ok": True,
                "data": rows,
                "error": None,
                "meta": {"asin": "B0TEST0000", "model": "Slate 7"},
            }

        monkeypatch.setattr(webapp_tools, "_api_query_trends", _fake)

        result = asyncio.run(
            dispatch_tool("query_trends", {"product": "Slate 7", "marketplace": "UK"})
        )

        assert result["ok"] is True
        assert isinstance(result["data"], dict), (
            "Phase 3 contract: query_trends must return a summary dict, not a row list"
        )
        assert result["data"]["count"] == 87
        assert "date_range" in result["data"]
        assert result["data"]["file_attached"].endswith(".xlsx")
        assert result["meta"]["asin"] == "B0TEST0000"
        # Preview is capped at MAX_PREVIEW_ROWS (3) — never the full row set.
        assert len(result["data"].get("preview", [])) <= 3

    def test_full_rows_land_in_xlsx_not_trimmed(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Excel must carry full DB fields, not the LLM-safe trimmed set."""
        from io import BytesIO

        from openpyxl import load_workbook

        _set_fake_env(monkeypatch)
        store = self._patch_session_and_step(monkeypatch)
        _reset_webapp_modules()
        from webapp import tools as webapp_tools
        from webapp.tools import dispatch_tool

        wide_row = {
            "id": 42,
            "site": "UK",
            "brand": "ExampleBrand",
            "model": "XR-100",
            "asin": "B0TESTTEST",
            "title": "MUST APPEAR IN XLSX",
            "price_cents": 14999,
            "url": "https://example.test",
        }

        def _fake(**_kw) -> dict:  # type: ignore[no-untyped-def]
            return {"ok": True, "data": [wide_row], "error": None, "meta": {}}

        monkeypatch.setattr(webapp_tools, "_api_query_latest", _fake)
        asyncio.run(dispatch_tool("query_latest", {"marketplace": "UK"}))

        pending = store.get("pending_files", [])
        assert len(pending) == 1, "exactly one xlsx should have been attached"
        xlsx_bytes = pending[0].content
        assert isinstance(xlsx_bytes, bytes)
        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws is not None
        headers = [c.value for c in ws[1]]
        for field in ("title", "url", "id"):
            assert field in headers, (
                f"{field!r} missing from xlsx — trim decorator is leaking "
                f"into the full-rows attachment path"
            )

    def test_query_log_appended_per_call(self, monkeypatch: pytest.MonkeyPatch) -> None:
        _set_fake_env(monkeypatch)
        store = self._patch_session_and_step(monkeypatch)
        _reset_webapp_modules()
        from webapp import tools as webapp_tools
        from webapp.tools import dispatch_tool

        def _fake(**_kw) -> dict:  # type: ignore[no-untyped-def]
            return {"ok": True, "data": [{"x": 1}], "error": None, "meta": {}}

        monkeypatch.setattr(webapp_tools, "_api_query_latest", _fake)

        asyncio.run(dispatch_tool("query_latest", {"marketplace": "UK"}))
        asyncio.run(dispatch_tool("query_latest", {"marketplace": "DE"}))

        log = store.get("query_log", [])
        assert len(log) == 2
        assert log[0]["tool"] == "query_latest"
        assert log[0]["args"]["marketplace"] == "UK"
        assert log[1]["args"]["marketplace"] == "DE"
        assert all("ts" in entry for entry in log)

    def test_failure_envelope_passes_through(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Failure envelopes (``ok=False``) must skip summarization entirely:
        no xlsx attach, no query_log append, so the LLM can see the full
        ``error`` string and the user is not offered a stale/empty download."""
        _set_fake_env(monkeypatch)
        store = self._patch_session_and_step(monkeypatch)
        _reset_webapp_modules()
        from webapp import tools as webapp_tools
        from webapp.tools import dispatch_tool

        failure = {
            "ok": False,
            "data": [],
            "error": "synthetic failure for test",
            "meta": {},
        }

        def _fake(**_kw) -> dict:  # type: ignore[no-untyped-def]
            return failure

        monkeypatch.setattr(webapp_tools, "_api_query_latest", _fake)
        result = asyncio.run(dispatch_tool("query_latest", {"marketplace": "UK"}))

        assert result is failure
        assert store.get("pending_files", []) == []
        assert store.get("query_log", []) == []

    def test_query_deals_summary_has_no_date_range(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """``query_deals`` uses ``date_field=None`` because ``start_time`` /
        ``end_time`` are Keepa-encoded minute integers; treating them as dates
        would produce a garbage range like ``'7584000 to 7590000'``. Until a
        decoder lands, the summary should simply omit ``date_range``."""
        _set_fake_env(monkeypatch)
        self._patch_session_and_step(monkeypatch)
        _reset_webapp_modules()
        from webapp import tools as webapp_tools
        from webapp.tools import dispatch_tool

        deal_row = {
            "asin": "B0TESTTEST",
            "site": "UK",
            "deal_type": "LIGHTNING",
            "start_time": 7584000,
            "end_time": 7590000,
        }

        def _fake(**_kw) -> dict:  # type: ignore[no-untyped-def]
            return {"ok": True, "data": [deal_row], "error": None, "meta": {}}

        monkeypatch.setattr(webapp_tools, "_api_query_deals", _fake)
        result = asyncio.run(dispatch_tool("query_deals", {"marketplace": "UK"}))

        assert result["ok"] is True
        assert isinstance(result["data"], dict)
        assert "date_range" not in result["data"], (
            "date_range on query_deals would stringify Keepa-encoded "
            "minute integers as dates; the decorator must skip it"
        )

    def test_tool_schema_size_stays_within_regression_budget(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Regression guard, not a compression-target test.

        Phase 2 baseline was ~6,500 chars. Phase 3 target is ≤5,500. Budget
        was raised from 6,000 → 8,000 when the web_search server-side schema
        and ``register_asin_from_url`` client schema landed (adds ~1,200
        chars). If this fires past the 8,000 bound, someone expanded a tool
        docstring — not the tool count.
        """
        import json

        _set_fake_env(monkeypatch)
        _reset_webapp_modules()
        from webapp.tools import TOOL_SCHEMAS

        size = len(json.dumps(TOOL_SCHEMAS, ensure_ascii=False))
        assert size <= 8000, (
            f"TOOL_SCHEMAS size {size} chars exceeds 8,000 regression budget "
            f"(bumped from 6,000 when web_search + register_asin_from_url "
            f"were added; Phase 2 baseline ~6,500)"
        )

    def test_attach_failure_drops_file_attached_from_summary(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """H1 regression: if the xlsx attach fails, the summary MUST NOT
        carry ``file_attached``. Otherwise the LLM tells the user a download
        is available that was never produced.
        """
        _set_fake_env(monkeypatch)
        self._patch_session_and_step(monkeypatch)
        _reset_webapp_modules()
        from webapp import summaries as webapp_summaries
        from webapp import tools as webapp_tools
        from webapp.tools import dispatch_tool

        def _fake(**_kw) -> dict:  # type: ignore[no-untyped-def]
            return {"ok": True, "data": [{"x": 1}], "error": None, "meta": {}}

        def _fail(*_args, **_kw) -> bool:  # type: ignore[no-untyped-def]
            return False

        monkeypatch.setattr(webapp_tools, "_api_query_latest", _fake)
        # Simulate attach failure by patching the helper itself — covers
        # both cl.File construction errors and session update errors.
        monkeypatch.setattr(webapp_summaries, "_attach_file_to_session", _fail)

        result = asyncio.run(dispatch_tool("query_latest", {"marketplace": "UK"}))

        assert result["ok"] is True
        data = result["data"]
        assert isinstance(data, dict)
        assert "file_attached" not in data, (
            "LLM must not see file_attached when the xlsx never reached the user"
        )
        assert data["file_attach_failed"] is True

    def test_xlsx_row_limit_truncates_and_flags_summary(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """M1 regression: a row list exceeding ``MAX_XLSX_ROWS`` must be
        truncated to the cap and the summary must carry ``xlsx_truncated``
        so the LLM can tell the user the attachment is incomplete.
        """
        from io import BytesIO

        from openpyxl import load_workbook

        _set_fake_env(monkeypatch)
        store = self._patch_session_and_step(monkeypatch)
        _reset_webapp_modules()
        from webapp import summaries as webapp_summaries
        from webapp import tools as webapp_tools
        from webapp.tools import dispatch_tool

        # Shrink the cap for a fast test — 5 rows out of 12.
        monkeypatch.setattr(webapp_summaries, "MAX_XLSX_ROWS", 5)
        rows = [{"i": i, "date": f"2026-04-{i:02d}"} for i in range(1, 13)]

        def _fake(**_kw) -> dict:  # type: ignore[no-untyped-def]
            return {"ok": True, "data": rows, "error": None, "meta": {}}

        monkeypatch.setattr(webapp_tools, "_api_query_latest", _fake)

        result = asyncio.run(dispatch_tool("query_latest", {"marketplace": "UK"}))

        data = result["data"]
        assert isinstance(data, dict)
        assert data["count"] == 12, "count reflects full DB rowset"
        assert data["xlsx_truncated"] is True
        assert data["xlsx_row_limit"] == 5
        pending = store.get("pending_files", [])
        wb = load_workbook(BytesIO(pending[0].content))
        ws = wb.active
        assert ws is not None
        # Header row + 5 data rows == 6 total.
        assert ws.max_row == 6, (
            f"xlsx must cap at MAX_XLSX_ROWS data rows; got {ws.max_row - 1}"
        )

    def test_warnings_are_truncated_before_reaching_llm(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """M3 regression: ``meta['warnings']`` is cap-limited (count + per-
        entry length) so a misbehaving upstream cannot silently 10x the
        summary token cost.
        """
        _set_fake_env(monkeypatch)
        self._patch_session_and_step(monkeypatch)
        _reset_webapp_modules()
        from webapp import summaries as webapp_summaries
        from webapp import tools as webapp_tools
        from webapp.tools import dispatch_tool

        long_msg = "X" * (webapp_summaries.MAX_WARNING_CHARS + 500)
        many_warnings = [long_msg] * (webapp_summaries.MAX_WARNINGS + 4)

        def _fake(**_kw) -> dict:  # type: ignore[no-untyped-def]
            return {
                "ok": True,
                "data": [{"x": 1}],
                "error": None,
                "meta": {"warnings": many_warnings},
            }

        monkeypatch.setattr(webapp_tools, "_api_query_latest", _fake)

        result = asyncio.run(dispatch_tool("query_latest", {"marketplace": "UK"}))

        data = result["data"]
        assert isinstance(data, dict)
        warnings = data["warnings"]
        # MAX_WARNINGS entries + one "more truncated" pointer line.
        assert len(warnings) == webapp_summaries.MAX_WARNINGS + 1
        assert all(
            len(w) <= webapp_summaries.MAX_WARNING_CHARS + 1
            for w in warnings[: webapp_summaries.MAX_WARNINGS]
        ), "each warning entry must be clipped to MAX_WARNING_CHARS (+ellipsis)"
        assert "more warnings truncated" in warnings[-1]

    def test_non_list_data_falls_back_without_crashing(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Defensive: if upstream leaks a dict where a list was promised,
        the decorator must log and return ``count=0`` instead of silently
        treating the dict as empty rows via ``or []``.
        """
        _set_fake_env(monkeypatch)
        self._patch_session_and_step(monkeypatch)
        _reset_webapp_modules()
        from webapp import tools as webapp_tools
        from webapp.tools import dispatch_tool

        def _fake(**_kw) -> dict:  # type: ignore[no-untyped-def]
            return {
                "ok": True,
                "data": {"unexpected": "shape"},
                "error": None,
                "meta": {},
            }

        monkeypatch.setattr(webapp_tools, "_api_query_latest", _fake)
        result = asyncio.run(dispatch_tool("query_latest", {"marketplace": "UK"}))

        assert result["ok"] is True
        data = result["data"]
        assert isinstance(data, dict)
        assert data["count"] == 0
        assert "preview" not in data


@pytest.mark.unit
class TestCacheControlWiring:
    """Verify run_chat_turn attaches cache_control to the last tool_result.

    Drives ``webapp.llm.run_chat_turn`` through a single tool round-trip using
    a stubbed Anthropic client, then inspects the ``history`` list to confirm
    the moving cache_control breakpoint is set. This is the core token-burn
    mitigation: every prior turn's prompt prefix must be cached for the next
    turn to pay ~10% instead of 100% of normal input cost.
    """

    def test_last_tool_result_block_gets_cache_control(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        _set_fake_env(monkeypatch)
        _reset_webapp_modules()
        from webapp import llm as webapp_llm

        class _Block:
            def __init__(self, **kw):
                self.__dict__.update(kw)

            def model_dump(self) -> dict:
                return dict(self.__dict__)

        class _Usage:
            def model_dump(self) -> dict:
                return {
                    "input_tokens": 10,
                    "output_tokens": 5,
                    "cache_creation_input_tokens": 0,
                    "cache_read_input_tokens": 0,
                }

        class _Resp:
            def __init__(self, content: list, stop_reason: str):
                self.content = content
                self.stop_reason = stop_reason
                self.usage = _Usage()

        tool_use_block = _Block(
            type="tool_use",
            id="toolu_X",
            name="keepa_budget",
            input={},
        )
        text_block = _Block(type="text", text="done")
        responses = iter(
            [
                _Resp([tool_use_block], "tool_use"),
                _Resp([text_block], "end_turn"),
            ]
        )

        def _fake_create(**_kwargs):
            return next(responses)

        monkeypatch.setattr(webapp_llm._client.messages, "create", _fake_create)

        async def _fake_dispatch(_name: str, _args: dict) -> dict:
            return {"ok": True, "data": [], "error": None, "meta": {}}

        monkeypatch.setattr(webapp_llm, "dispatch_tool", _fake_dispatch)

        history: list[dict] = [{"role": "user", "content": "what's the keepa budget?"}]
        final_text, updated = asyncio.run(webapp_llm.run_chat_turn(history))

        assert final_text == "done"

        tool_result_msgs = [
            m
            for m in updated
            if m["role"] == "user"
            and isinstance(m["content"], list)
            and any(b.get("type") == "tool_result" for b in m["content"])
        ]
        assert len(tool_result_msgs) == 1, (
            "Expected exactly one user message carrying tool_result blocks"
        )

        last_block = tool_result_msgs[0]["content"][-1]
        assert last_block.get("cache_control") == {"type": "ephemeral"}, (
            "Last tool_result block must be marked ephemeral for moving breakpoint caching"
        )

    def test_cache_control_does_not_accumulate_across_turns(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """After N sequential user turns that each invoke a tool, the history
        must still contain exactly ONE tool_result block with cache_control.

        Regression guard for the production 400: "A maximum of 4 blocks with
        cache_control may be provided. Found 5." The moving cache_control
        marker must actually MOVE (strip old, add new), not accumulate.
        """
        _set_fake_env(monkeypatch)
        _reset_webapp_modules()
        from webapp import llm as webapp_llm

        class _Block:
            def __init__(self, **kw):
                self.__dict__.update(kw)

            def model_dump(self) -> dict:
                return dict(self.__dict__)

        class _Usage:
            def model_dump(self) -> dict:
                return {}

        class _Resp:
            def __init__(self, content: list, stop_reason: str):
                self.content = content
                self.stop_reason = stop_reason
                self.usage = _Usage()

        def _tool_use(idx: int) -> _Block:
            return _Block(
                type="tool_use",
                id=f"toolu_{idx}",
                name="keepa_budget",
                input={},
            )

        text_block = _Block(type="text", text="done")

        # Five user turns in a row, each doing exactly one tool_use then
        # resolving to an end_turn. That's 10 create() calls total.
        responses = iter(
            [
                _Resp([_tool_use(1)], "tool_use"),
                _Resp([text_block], "end_turn"),
                _Resp([_tool_use(2)], "tool_use"),
                _Resp([text_block], "end_turn"),
                _Resp([_tool_use(3)], "tool_use"),
                _Resp([text_block], "end_turn"),
                _Resp([_tool_use(4)], "tool_use"),
                _Resp([text_block], "end_turn"),
                _Resp([_tool_use(5)], "tool_use"),
                _Resp([text_block], "end_turn"),
            ]
        )

        def _fake_create(**_kwargs):
            return next(responses)

        monkeypatch.setattr(webapp_llm._client.messages, "create", _fake_create)

        async def _fake_dispatch(_name: str, _args: dict) -> dict:
            return {"ok": True, "data": [], "error": None, "meta": {}}

        monkeypatch.setattr(webapp_llm, "dispatch_tool", _fake_dispatch)

        history: list[dict] = []

        async def _drive_five_turns() -> list[dict]:
            for turn_i in range(5):
                history.append({"role": "user", "content": f"turn {turn_i}"})
                _, hist_after = await webapp_llm.run_chat_turn(history)
                # run_chat_turn mutates the same list in place, so just
                # continue with the updated reference.
                assert hist_after is history
            return history

        asyncio.run(_drive_five_turns())

        # Count every block in history that carries cache_control.
        marked_blocks = [
            blk
            for msg in history
            if msg["role"] == "user" and isinstance(msg["content"], list)
            for blk in msg["content"]
            if isinstance(blk, dict) and blk.get("cache_control") is not None
        ]
        assert len(marked_blocks) == 1, (
            f"Expected exactly 1 tool_result with cache_control after 5 turns, "
            f"found {len(marked_blocks)}. This means the moving breakpoint is "
            f"accumulating instead of moving, and production will hit the "
            f"Anthropic 4-block limit."
        )

        # It must be on a tool_result block (not an assistant tool_use etc.)
        assert marked_blocks[0].get("type") == "tool_result"

        # And it must be on the LATEST tool_result — verify by scanning in
        # reverse order for the first tool_result block.
        last_tool_result = None
        for msg in reversed(history):
            if msg["role"] == "user" and isinstance(msg["content"], list):
                for blk in msg["content"]:
                    if isinstance(blk, dict) and blk.get("type") == "tool_result":
                        last_tool_result = blk
                        break
            if last_tool_result is not None:
                break

        assert last_tool_result is marked_blocks[0], (
            "cache_control must be on the LATEST tool_result block"
        )

    def test_end_turn_without_tool_use_does_not_crash(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Safety rail: the cache_control branch must not fire when the
        iteration produces no tool_results (e.g., an end_turn reached on the
        first pass). Otherwise the list-index access would IndexError."""
        _set_fake_env(monkeypatch)
        _reset_webapp_modules()
        from webapp import llm as webapp_llm

        class _Block:
            def __init__(self, **kw):
                self.__dict__.update(kw)

            def model_dump(self) -> dict:
                return dict(self.__dict__)

        class _Usage:
            def model_dump(self) -> dict:
                return {}

        class _Resp:
            def __init__(self, content: list, stop_reason: str):
                self.content = content
                self.stop_reason = stop_reason
                self.usage = _Usage()

        text_block = _Block(type="text", text="hello back")

        def _fake_create(**_kwargs):
            return _Resp([text_block], "end_turn")

        monkeypatch.setattr(webapp_llm._client.messages, "create", _fake_create)

        final_text, updated = asyncio.run(
            webapp_llm.run_chat_turn([{"role": "user", "content": "hi"}])
        )

        assert final_text == "hello back"
        assert not any(
            m["role"] == "user"
            and isinstance(m["content"], list)
            and any(b.get("type") == "tool_result" for b in m["content"])
            for m in updated
        )


@pytest.mark.unit
class TestAsyncThreadOffload:
    """Bug A (issue #13): async wrappers must offload blocking sync I/O
    so one user's slow Keepa fetch does not freeze the Chainlit event loop
    for every other concurrent session.
    """

    def test_every_step_wrapper_uses_asyncio_to_thread(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        import threading

        _set_fake_env(monkeypatch)

        import chainlit as cl

        def _noop_step(**_kwargs):  # type: ignore[no-untyped-def]
            def _decorator(fn):  # type: ignore[no-untyped-def]
                return fn

            return _decorator

        monkeypatch.setattr(cl, "step", _noop_step)

        _reset_webapp_modules()
        from webapp import tools as webapp_tools
        from webapp.tools import TOOL_SCHEMAS, dispatch_tool

        def _record_thread(**_kw) -> dict:  # type: ignore[no-untyped-def]
            return {
                "ok": True,
                "data": [],
                "error": None,
                "meta": {"_thread": threading.current_thread().name},
            }

        for attr in (
            "_api_check_freshness",
            "_api_keepa_budget",
            "_api_query_availability",
            "_api_query_compare",
            "_api_query_deals",
            "_api_query_latest",
            "_api_query_ranking",
            "_api_query_sellers",
            "_api_query_trends",
            "_api_register_asin_from_url",
        ):
            monkeypatch.setattr(webapp_tools, attr, _record_thread)

        async def _run_all() -> list[tuple[str, dict]]:
            out: list[tuple[str, dict]] = []
            for tool in TOOL_SCHEMAS:
                # Skip server-side tools (no input_schema → Anthropic-dispatched).
                if "input_schema" not in tool:
                    continue
                name = tool["name"]
                args: dict = {}
                for prop in tool["input_schema"].get("required", []):
                    args[prop] = "UK" if prop == "marketplace" else "Slate 7"
                out.append((name, await dispatch_tool(name, args)))
            return out

        for name, result in asyncio.run(_run_all()):
            thread_name = result["meta"]["_thread"]
            assert not thread_name.startswith("MainThread"), (
                f"{name}: sync _api_* ran on {thread_name!r} — Bug A regression. "
                f"Wrapper must await asyncio.to_thread(_api_*, ...)."
            )


@pytest.mark.unit
class TestWebSearchTool:
    """Server-side web_search schema shape + anti-code-execution guard."""

    def test_web_search_tool_declared(self, monkeypatch: pytest.MonkeyPatch) -> None:
        _set_fake_env(monkeypatch)
        _reset_webapp_modules()
        from webapp.tools import TOOL_SCHEMAS

        web_tools = [t for t in TOOL_SCHEMAS if t.get("type") == "web_search_20260209"]
        assert len(web_tools) == 1, (
            "Expected exactly one web_search_20260209 server-side tool declaration"
        )
        tool = web_tools[0]
        assert tool["name"] == "web_search", (
            "Anthropic requires the name literal 'web_search'"
        )
        allowed = set(tool.get("allowed_domains", []))
        for required in ("amazon.com", "amazon.de", "amazon.co.uk", "amazon.co.jp"):
            assert required in allowed, (
                f"allowed_domains must cover {required} — missing means "
                f"web_search can't reach that marketplace"
            )

    def test_no_code_execution_tool_declared(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """Decision regression guard.

        ``web_search_20260209`` has dynamic filtering built-in — Anthropic
        provisions a code-execution sandbox internally. Declaring a
        standalone ``code_execution_*`` tool alongside creates a second
        environment and confuses the model. This test blocks that.
        """
        _set_fake_env(monkeypatch)
        _reset_webapp_modules()
        from webapp.tools import TOOL_SCHEMAS

        code_tools = [
            t for t in TOOL_SCHEMAS
            if str(t.get("type") or "").startswith("code_execution")
        ]
        assert code_tools == [], (
            f"Must NOT declare a standalone code_execution tool alongside "
            f"web_search_20260209; found: {[t.get('type') for t in code_tools]}"
        )


@pytest.mark.unit
class TestRegisterAsinFromUrlDispatch:
    """register_asin_from_url routes through dispatch_tool correctly and
    bypasses the summarize_for_llm decorator (returns a small dict, not rows).
    """

    def test_register_asin_from_url_skips_summarize_decorator(
        self, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Success response is a small dict envelope; no xlsx / pending_files.

        The summarize decorator is wired to row-emitting query tools only.
        register_asin_from_url's data is a compact dict with
        asin/marketplace/product_id — summary would mis-handle it as 0 rows
        and attach a useless empty xlsx. This test ensures the dispatch
        path never reaches summarize for this tool.
        """
        _set_fake_env(monkeypatch)

        from dataclasses import dataclass

        import chainlit as cl

        store: dict = {}

        class _FakeSession:
            def get(self, k, default=None):
                return store.get(k, default)

            def set(self, k, v):
                store[k] = v

        @dataclass
        class _FakeFile:
            name: str = ""
            content: bytes | str | None = None
            mime: str | None = None
            display: str = "inline"

        def _noop_step(**_kwargs):  # type: ignore[no-untyped-def]
            def _decorator(fn):  # type: ignore[no-untyped-def]
                return fn

            return _decorator

        monkeypatch.setattr(cl, "user_session", _FakeSession())
        monkeypatch.setattr(cl, "step", _noop_step)
        monkeypatch.setattr(cl, "File", _FakeFile)

        _reset_webapp_modules()
        from webapp import tools as webapp_tools
        from webapp.tools import dispatch_tool

        def _fake(**_kw) -> dict:  # type: ignore[no-untyped-def]
            return {
                "ok": True,
                "data": {
                    "asin": "B0TESTTEST1",
                    "marketplace": "DE",
                    "brand": "TP-Link",
                    "model": "AX1500",
                    "product_id": 42,
                    "registered": True,
                    "new_product": True,
                },
                "error": None,
                "meta": {},
            }

        monkeypatch.setattr(webapp_tools, "_api_register_asin_from_url", _fake)

        result = asyncio.run(
            dispatch_tool(
                "register_asin_from_url",
                {
                    "brand": "TP-Link",
                    "model": "AX1500",
                    "marketplace": "DE",
                    "amazon_url": "https://www.amazon.de/dp/B0TESTTEST1",
                },
            )
        )

        assert result["ok"] is True
        assert result["data"]["asin"] == "B0TESTTEST1"
        assert store.get("pending_files", []) == [], (
            "register_asin_from_url returns a compact dict; no xlsx "
            "attachment should be created"
        )


@pytest.mark.unit
class TestPauseTurnHandling:
    """Server-side tools (web_search) can hit Anthropic's 10-iter server loop
    cap, returning ``stop_reason='pause_turn'``. webapp/llm.py must resume
    transparently by re-sending the same messages, WITHOUT injecting a fake
    'Continue' user message.
    """

    def test_pause_turn_resume(self, monkeypatch: pytest.MonkeyPatch) -> None:
        _set_fake_env(monkeypatch)
        _reset_webapp_modules()
        from webapp import llm as webapp_llm

        class _Block:
            def __init__(self, **kw):
                self.__dict__.update(kw)

            def model_dump(self) -> dict:
                return dict(self.__dict__)

        class _Usage:
            def model_dump(self) -> dict:
                return {}

        class _Resp:
            def __init__(self, content: list, stop_reason: str):
                self.content = content
                self.stop_reason = stop_reason
                self.usage = _Usage()

        # First response: server tool pause (no client tool_use; trailing
        # server_tool_use is the resume signal Anthropic expects).
        server_tool_use_block = _Block(
            type="server_tool_use",
            id="srv_abc",
            name="web_search",
            input={"query": "site:amazon.de slate 7"},
        )
        text_final = _Block(type="text", text="found it")

        call_count = {"n": 0}
        responses = [
            _Resp([server_tool_use_block], "pause_turn"),
            _Resp([text_final], "end_turn"),
        ]

        def _fake_create(**kwargs):
            call_count["n"] += 1
            idx = call_count["n"] - 1
            assert idx < len(responses), "messages.create invoked too many times"
            return responses[idx]

        monkeypatch.setattr(webapp_llm._client.messages, "create", _fake_create)

        history: list[dict] = [{"role": "user", "content": "find slate 7 in DE"}]
        final_text, updated = asyncio.run(webapp_llm.run_chat_turn(history))

        assert final_text == "found it"
        assert call_count["n"] == 2, (
            "Expected exactly 2 calls to messages.create: initial + resume. "
            f"Got {call_count['n']}."
        )

        # Guard: no fake "Continue" user message was injected between the
        # pause_turn assistant turn and the resumed call. The only user
        # entries should be the original prompt.
        user_msgs = [m for m in updated if m.get("role") == "user"]
        assert len(user_msgs) == 1, (
            f"Expected exactly 1 user message (the original prompt); "
            f"got {len(user_msgs)} — suggests a fake 'Continue' was injected"
        )
        assert user_msgs[0]["content"] == "find slate 7 in DE"


@pytest.mark.unit
class TestBlockCountMonitor:
    """The 20-block lookback limit on prompt-cache breakpoints means long
    turns (web_search + register + re-query) can silently cache-miss.
    webapp/llm.py emits a warning once history exceeds 15 blocks so ops
    can catch drift.
    """

    def test_history_block_count_warning_logged(
        self, monkeypatch: pytest.MonkeyPatch, caplog: pytest.LogCaptureFixture
    ) -> None:
        _set_fake_env(monkeypatch)
        _reset_webapp_modules()
        from webapp import llm as webapp_llm

        class _Block:
            def __init__(self, **kw):
                self.__dict__.update(kw)

            def model_dump(self) -> dict:
                return dict(self.__dict__)

        class _Usage:
            def model_dump(self) -> dict:
                return {}

        class _Resp:
            def __init__(self, content: list, stop_reason: str):
                self.content = content
                self.stop_reason = stop_reason
                self.usage = _Usage()

        tool_use = _Block(type="tool_use", id="toolu_X", name="keepa_budget", input={})
        text = _Block(type="text", text="done")

        responses = iter([_Resp([tool_use], "tool_use"), _Resp([text], "end_turn")])
        monkeypatch.setattr(
            webapp_llm._client.messages, "create", lambda **_k: next(responses)
        )

        async def _fake_dispatch(_name: str, _args: dict) -> dict:
            return {"ok": True, "data": [], "error": None, "meta": {}}

        monkeypatch.setattr(webapp_llm, "dispatch_tool", _fake_dispatch)

        # Prime history with > 15 content blocks so the warning fires on the
        # first tool_use round. A single user message with 16 text blocks
        # gets us over the threshold.
        primer_content = [
            {"type": "text", "text": f"block {i}"} for i in range(16)
        ]
        history: list[dict] = [{"role": "user", "content": primer_content}]

        with caplog.at_level("WARNING", logger=webapp_llm.logger.name):
            asyncio.run(webapp_llm.run_chat_turn(history))

        warnings = [r for r in caplog.records if "total_blocks=" in r.getMessage()]
        assert warnings, (
            "Expected a warning containing 'total_blocks=' when history "
            "exceeds the 15-block threshold"
        )
        assert any(
            "20-block cache lookback" in r.getMessage() for r in warnings
        ), "Warning should name the 20-block lookback limit for ops clarity"


@pytest.mark.unit
class TestSystemPromptContent:
    """SYSTEM_PROMPT must include the ASIN Discovery Flow so the LLM knows
    when to call web_search and register_asin_from_url without being told
    on every turn.
    """

    def test_system_prompt_contains_asin_flow(self, monkeypatch: pytest.MonkeyPatch) -> None:
        _set_fake_env(monkeypatch)
        _reset_webapp_modules()
        from webapp.config import SYSTEM_PROMPT

        assert "register_asin_from_url" in SYSTEM_PROMPT
        assert "web_search" in SYSTEM_PROMPT
        assert "ASIN Discovery Flow" in SYSTEM_PROMPT
        # Prompt-injection defense lines must be present.
        assert "do not fabricate Amazon URLs" in SYSTEM_PROMPT

